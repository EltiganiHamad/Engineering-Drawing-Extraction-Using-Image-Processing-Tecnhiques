# -*- coding: utf-8 -*-
"""
Created on Wed Nov 25 21:00:17 2020

@author: eltig
"""

import cv2
import numpy as np
import matplotlib.pyplot as plt
import pytesseract
import openpyxl
import os


pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'


def sort_contours(cnts, method="left-to-right"):
	# initialize sort index
	reverse = False
	i = 0
	# This is if we use the reverse approach
	if method == "right-to-left" or method == "bottom-to-top":
		reverse = True
	# sorting of Y coordinates rather than x coordinates
	if method == "top-to-bottom" or method == "bottom-to-top":
		i = 1
	# sorting top to bottom by constructing bounding boxes
	boundingBoxes = [cv2.boundingRect(c) for c in cnts]
	(cnts, boundingBoxes) = zip(*sorted(zip(cnts, boundingBoxes),
		key=lambda b:b[1][i], reverse=reverse))
	# calling all the list of contours and bounding boxes
	return (cnts, boundingBoxes)

def diagram_segmentation(img_path):
    # Read image
    img = cv2.imread(img_path,0)
         
    # Image thresholding
    (thresh, img_bin) = cv2.threshold(img, 120, 255,cv2.THRESH_BINARY|cv2.THRESH_OTSU)
            
    # Converting image to binary
    img_bin = 255-img_bin 
    cv2.imwrite("Image_bin.jpg",img_bin)
        
    # Defining a kernel length
    kernel_length = np.array(img).shape[1]//50
             
    # Verticle kernel to detect all the vertical lines in the image
    verticle_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, kernel_length))
    # Horizontal kernel to detect all the horizontal lines in the image
    hori_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_length, 1))
    # Set up a (3 x 3) kernel
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
            
    # Detect verticle lines using morphological operations
    img_temp1 = cv2.erode(img_bin, verticle_kernel, iterations=3)
    verticle_lines_img = cv2.dilate(img_temp1, verticle_kernel, iterations=3)
    cv2.imwrite("verticle_lines.jpg",verticle_lines_img)
            
    # Detect horizontal lines using morphological operations
    img_temp2 = cv2.erode(img_bin, hori_kernel, iterations=3)
    horizontal_lines_img = cv2.dilate(img_temp2, hori_kernel, iterations=3)
    cv2.imwrite("horizontal_lines.jpg",horizontal_lines_img)
            
    # To add into the new image weighting parameters were used
    alpha = 0.5
    beta = 1.0 - alpha
    # Add the two images to get a third image
    img_final_bin = cv2.addWeighted(verticle_lines_img, alpha, horizontal_lines_img, beta, 0.0)
    img_final_bin = cv2.erode(~img_final_bin, kernel, iterations=2)
    (thresh, img_final_bin) = cv2.threshold(img_final_bin, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
            
    # Debugging
    # The verticle and horizantal lines combination is tested out
    cv2.imwrite("img_final_bin.jpg",img_final_bin)
    
    # Find contours for image, which will detect all the relevent region of interest
    contours, hierarchy = cv2.findContours(img_final_bin, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    # Sort all the contours by top to bottom approach
    (contours, boundingBoxes) = sort_contours(contours, method="top-to-bottom")
    
    # Find the size of the image
    [nrow,ncol] = img.shape
    # Apply the masks
    drawing_mask = np.ones((nrow,ncol)) 
    table_mask = np.zeros((nrow,ncol)) 
    for c in contours:
        # To find the location and width and height of the contours
        x, y, w, h = cv2.boundingRect(c)
    
        # If width is between 80 to 1800 and Height between 20 to 1500 except line 286 and 287. Apply mask!
        if (80 < w < 1800 and 20 < h < 1500 ) and w > h and w !=(286) and w !=(287):
            drawing_mask[y:y+h, x:x+w] = 0
            table_mask[y:y+h, x:x+w] = 1
        
    drawing = img * drawing_mask
   
    tables = img * table_mask
    plt.imshow(drawing,cmap='gray')
    cv2.imwrite('Drawing.png', drawing)
    cv2.imwrite('Tables.png', tables)
    

file =  r'Tables.png'
    
def extractTabularData1(file):
    
    #Creating a new workshop using the workbook class and titling "AMMENDMENTS"
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "AMMENDMENTS"
 

    
    #Reading in the image
    im1 = cv2.imread(file, 0)
    im = cv2.imread(file)

    #Bluring the image
    blur = cv2.blur(im1,(1,1))

    #Thresholding the image
    ret,thresh_value = cv2.threshold(blur,100,255,cv2.THRESH_BINARY_INV)

 
    #Initializing the kernel for image erosion
    kernel1 = np.ones((3,3), np.uint8)
    
    #Eroding the image
    img_erosion = cv2.erode(thresh_value, kernel1, iterations=1)

    #Initializing the kernel for image dialation
    kernel2 = np.ones((3,3),np.uint8)
    
    #Dialating the image
    dilated_value = cv2.dilate(img_erosion,kernel2,iterations = 1)

    #Finding the Contours within the image
    contours, hierarchy = cv2.findContours(dilated_value,cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
    
    #Initializing a list to which the contour cordinates will be appended
    cordinates = []
    
            
    #For every countour from the variable counters
    for cnt in contours:
        #Bounds a rectangle around the countour
        x,y,w,h = cv2.boundingRect(cnt)
        #appending the cordinates of the rectangle to the list
        cordinates.append((x,y,w,h))
        #Extracting the value of the countor
        cropped = im1[y:y + h, x:x + w] 
        #Opening a text file in append mode
        file = open("recognized.txt", "a")
        #Converting the extracted countour values into string using pytesseract
        text = pytesseract.image_to_string(cropped)
        #Writing the string values into the text files
        file.write(str(text)) 
        file.write("\n")
        #Closing the file
        file.close() 
        #Drawing the countours on our untouched image
        if y > 50:
            cv2.rectangle(im,(x,y),(x+w,y+h),(0,0,255),1)
   
    #Showing and Saving the Detected image
    plt.imshow(im)
    cv2.namedWindow('detecttable', cv2.WINDOW_NORMAL)
    cv2.imwrite('detecttable.jpg',im)
    #Opening the text file    
    f = open("recognized.txt", "r")
    #Reading each line from the text file
    content_list = f.readlines()
    #For each row of text in the variable
    for row in range(len(content_list)):
        #Filtering the unidentifiable text. 
        if(content_list[row] != '\x0c\n'):
         #append the lines to the exel sheet
         sheet.append([content_list[row]])
    #Closing the file
    f.close()
    
    #Extracting the data from the text file and saving it into another text file without spaces
    f1 = open("recognized.txt", "r")
    for line in f1:
        if not line.isspace():
            file2 = open("DIPimagedetailsWITHOUTspace.txt", "a")
            file2.write(line)
            file2.close()
    f1.close()        
            
    #Converting the file into an object        
    with open("DIPimagedetailsWITHOUTspace.txt", 'r') as read_obj:
            #Creating a new sheet
            newsheet =  wb.create_sheet("Extract", 0)
            
            #Searching for the field names in the text file and appending them into the sheet
            for line in read_obj:    

                string_to_search1 = ("UNIT:")
                string_to_search2 = ("DRAWN:")
                string_to_search3 = ("DRAWING TITLE:")
                string_to_search4 = ("CHECKED:")
                string_to_search5 = ("APPROVED:")
                string_to_search6 = ("STATUS:")
                string_to_search7 = ("COMPANY:")
                string_to_search8 = ("PAGE:")
                string_to_search9 = ("DRAWING NUMBER:")



                newsheet['A1'] = (string_to_search1)    
                newsheet['A2'] = (string_to_search2)
                newsheet['A3'] = (string_to_search3)
                newsheet['A4'] = (string_to_search4)
                newsheet['A5'] = (string_to_search5)
                newsheet['A6'] = (string_to_search6)
                newsheet['A7'] = (string_to_search7)
                newsheet['A8'] = (string_to_search8)
                newsheet['A9'] = (string_to_search9)

                #Appending the line below each field name and appending it into the sheet
                if string_to_search1 in line:
                    newsheet['B1'] = next(read_obj)

                if string_to_search2 in line:
                    newsheet['B2'] = next(read_obj)

                if string_to_search3 in line:
                    newsheet['B3'] = next(read_obj)      

                if string_to_search4 in line:
                    newsheet['B4'] = next(read_obj)

                if string_to_search5 in line:
                    newsheet['B5'] = next(read_obj) 

                if string_to_search6 in line:
                    newsheet['B6'] = next(read_obj)

                if string_to_search7 in line:
                    newsheet['B7'] = next(read_obj)

                if string_to_search8 in line:
                    newsheet['B8'] = next(read_obj)

                if string_to_search9 in line:
                    newsheet['B9'] = next(read_obj)


                #Saving the excel sheet
                wb.save("DIPassignment.xlsx")  
    #closing the text file object
    read_obj.close()
    #Deleting the text files
    os.remove("recognized.txt")
    os.remove("DIPimagedetailsWITHOUTspace.txt")
    
def extractTabularData2(file):
    wb = openpyxl.Workbook() 
    sheet = wb.active
    sheet.title = "AMMENDMENTS"
 

    for x in range(1,10):
        for y in range(1,10):
            sheet.cell(row=x, column=y)
    
    im1 = cv2.imread(file, 0)
    im = cv2.imread(file)

 

    blur = cv2.blur(im1,(1,1))

 


    ret,thresh_value = cv2.threshold(blur,100,255,cv2.THRESH_BINARY_INV)

 

    kernel1 = np.ones((3,3), np.uint8)
    img_erosion = cv2.erode(thresh_value, kernel1, iterations=1)

 


    kernel2 = np.ones((3,3),np.uint8)
    dilated_value = cv2.dilate(img_erosion,kernel2,iterations = 1)

 

    contours, hierarchy = cv2.findContours(dilated_value,cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
    cordinates = []
    
            

    for cnt in contours:
        x,y,w,h = cv2.boundingRect(cnt)
        cordinates.append((x,y,w,h))
        cropped = im1[y:y + h, x:x + w] 
        file = open("recognized.txt", "a")
        text = pytesseract.image_to_string(cropped)     
        file.write(str(text)) 
        file.write("\n")
        file.close() 
        #bounding the images
        if y > 50:
            cv2.rectangle(im,(x,y),(x+w,y+h),(0,0,255),1)
     
    
    plt.imshow(im)
    cv2.namedWindow('detecttable', cv2.WINDOW_NORMAL)
    cv2.imwrite('detecttable.jpg',im)    
    f = open("recognized.txt", "r")
    content_list = f.readlines()
    for row in range(len(content_list)):
        if(content_list[row] != '\x0c\n'):
         sheet.append([content_list[row]])
    f.close()
    
    
    f1 = open("recognized.txt", "r")
    for line in f1:
        if not line.isspace():
            file2 = open("DIPimagedetailsWITHOUTspace.txt", "a")
            file2.write(line)
            file2.close()
    f1.close()        
            
            
    with open("DIPimagedetailsWITHOUTspace.txt", 'r') as read_obj:
            newsheet =  wb.create_sheet("Extract", 0)
            

            for line in read_obj:    

                string_to_search1 = ("UNIT:")
                string_to_search2 = ("DRAWN BY:")
                string_to_search3 = ("TITLE:")
                string_to_search4 = ("CHECKED BY:")
                string_to_search5 = ("APPROVED BY:")
                string_to_search6 = ("STATUS:")
                string_to_search7 = ("COMPANY NAME:")
                string_to_search8 = ("PAGE:")
                string_to_search9 = ("DRAWING NO.:")



                newsheet['A1'] = (string_to_search1)    
                newsheet['A2'] = (string_to_search2)
                newsheet['A3'] = (string_to_search3)
                newsheet['A4'] = (string_to_search4)
                newsheet['A5'] = (string_to_search5)
                newsheet['A6'] = (string_to_search6)
                newsheet['A7'] = (string_to_search7)
                newsheet['A8'] = (string_to_search8)
                newsheet['A9'] = (string_to_search9)


                if string_to_search1 in line:
                    newsheet['B1'] = next(read_obj)

                if string_to_search2 in line:
                    newsheet['B2'] = next(read_obj)

                if string_to_search3 in line:
                    newsheet['B3'] = next(read_obj)      

                if string_to_search4 in line:
                    newsheet['B4'] = next(read_obj)

                if string_to_search5 in line:
                    newsheet['B5'] = next(read_obj) 

                if string_to_search6 in line:
                    newsheet['B6'] = next(read_obj)

                if string_to_search7 in line:
                    newsheet['B7'] = next(read_obj)

                if string_to_search8 in line:
                    newsheet['B8'] = next(read_obj)

                if string_to_search9 in line:
                    newsheet['B9'] = next(read_obj)



                wb.save("DIPassignment.xlsx")  
  
    read_obj.close()
    os.remove("recognized.txt")
    os.remove("DIPimagedetailsWITHOUTspace.txt")
    
def extractTabularData3(file):
    wb = openpyxl.Workbook() 
    sheet = wb.active
    sheet.title = "AMMENDMENTS"
 

    for x in range(1,10):
        for y in range(1,10):
            sheet.cell(row=x, column=y)
    
    im1 = cv2.imread(file, 0)
    im = cv2.imread(file)

 

    blur = cv2.blur(im1,(1,1))

 


    ret,thresh_value = cv2.threshold(blur,100,255,cv2.THRESH_BINARY_INV)

 

    kernel1 = np.ones((1,1), np.uint8)
    img_erosion = cv2.erode(thresh_value, kernel1, iterations=1)

 


    kernel2 = np.ones((4,4),np.uint8)
    dilated_value = cv2.dilate(img_erosion,kernel2,iterations = 1)

 

    contours, hierarchy = cv2.findContours(dilated_value,cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
    cordinates = []
   
            

    for cnt in contours:
        x,y,w,h = cv2.boundingRect(cnt)
        cordinates.append((x,y,w,h))
        cropped = im1[y:y + h, x:x + w] 
        file = open("recognized.txt", "a")
        text = pytesseract.image_to_string(cropped)
        file.write(str(text)) 
        file.write("\n")
        file.close() 
        #bounding the images
        if y > 50:
            cv2.rectangle(im,(x,y),(x+w,y+h),(0,0,255),1)
    
    plt.imshow(im)
    cv2.namedWindow('detecttable', cv2.WINDOW_NORMAL)
    cv2.imwrite('detecttable.jpg',im)
        
    f = open("recognized.txt", "r")
    content_list = f.readlines()
    for row in range(len(content_list)):
        if(content_list[row] != '\x0c\n'):
         sheet.append([content_list[row]])
    f.close()
    
    
    f1 = open("recognized.txt", "r")
    for line in f1:
        if not line.isspace():
            file2 = open("DIPimagedetailsWITHOUTspace.txt", "a")
            file2.write(line)
            file2.close()
    f1.close()        
            
            
    with open("DIPimagedetailsWITHOUTspace.txt", 'r') as read_obj:
            newsheet =  wb.create_sheet("Extract", 0)
            

            for line in read_obj:    

                string_to_search1 = ("UNIT:")
                string_to_search2 = ("DRAWN:")
                string_to_search3 = ("TITLE:")
                string_to_search4 = ("CHECKED:")
                string_to_search5 = ("APPROVED:")
                string_to_search6 = ("STS:")
                string_to_search7 = ("PROJECT NO:")
                string_to_search8 = ("PAGE:")
                string_to_search9 = ("DRAWING NO.:")



                newsheet['A1'] = (string_to_search1)    
                newsheet['A2'] = (string_to_search2)
                newsheet['A3'] = (string_to_search3)
                newsheet['A4'] = (string_to_search4)
                newsheet['A5'] = (string_to_search5)
                newsheet['A6'] = (string_to_search6)
                newsheet['A7'] = (string_to_search7)
                newsheet['A8'] = (string_to_search8)
                newsheet['A9'] = (string_to_search9)


                if string_to_search1 in line:
                    newsheet['B1'] = next(read_obj)

                if string_to_search2 in line:
                    newsheet['B2'] = next(read_obj)

                if string_to_search3 in line:
                    newsheet['B3'] = next(read_obj)      

                if string_to_search4 in line:
                    newsheet['B4'] = next(read_obj)

                if string_to_search5 in line:
                    newsheet['B5'] = next(read_obj) 

                if string_to_search6 in line:
                    newsheet['B6'] = next(read_obj)

                if string_to_search7 in line:
                    newsheet['B7'] = next(read_obj)

                if string_to_search8 in line:
                    newsheet['B8'] = next(read_obj)

                if string_to_search9 in line:
                    newsheet['B9'] = next(read_obj)



                wb.save("DIPassignment.xlsx")  
  
    read_obj.close()
    os.remove("recognized.txt")
    os.remove("DIPimagedetailsWITHOUTspace.txt")
    
def extractTabularData4(file):
    wb = openpyxl.Workbook() 
    sheet = wb.active
    sheet.title = "AMMENDMENTS"
 

    for x in range(1,10):
        for y in range(1,10):
            sheet.cell(row=x, column=y)
    
    im1 = cv2.imread(file, 0)
    im = cv2.imread(file)

 

    blur = cv2.blur(im1,(1,1))

 


    ret,thresh_value = cv2.threshold(blur,100,255,cv2.THRESH_BINARY_INV)

 

    kernel1 = np.ones((2,2), np.uint8)
    img_erosion = cv2.erode(thresh_value, kernel1, iterations=1)

 


    kernel2 = np.ones((2,2),np.uint8)
    dilated_value = cv2.dilate(img_erosion,kernel2,iterations = 1)

 

    contours, hierarchy = cv2.findContours(dilated_value,cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
    cordinates = []
    
            

    for cnt in contours:
        x,y,w,h = cv2.boundingRect(cnt)
        cordinates.append((x,y,w,h))
        cropped = im1[y:y + h, x:x + w] 
        file = open("recognized.txt", "a")
        text = pytesseract.image_to_string(cropped)
        file.write(str(text)) 
        file.write("\n")
        file.close() 
        #bounding the images
        if y > 50:
            cv2.rectangle(im,(x,y),(x+w,y+h),(0,0,255),1)
            
    plt.imshow(im)
    cv2.namedWindow('detecttable', cv2.WINDOW_NORMAL)
    cv2.imwrite('detecttable.jpg',im)
        
    f = open("recognized.txt", "r")
    content_list = f.readlines()
    for row in range(len(content_list)):
        if(content_list[row] != '\x0c\n'):
         sheet.append([content_list[row]])
    f.close()
    
    
    f1 = open("recognized.txt", "r")
    for line in f1:
        if not line.isspace():
            file2 = open("DIPimagedetailsWITHOUTspace.txt", "a")
            file2.write(line)
            file2.close()
    f1.close()        
            
            
    with open("DIPimagedetailsWITHOUTspace.txt", 'r') as read_obj:
            newsheet =  wb.create_sheet("Extract", 0)
            

            for line in read_obj:    

                string_to_search1 = ("UNIT")
                string_to_search2 = ("DRAWN BY")
                string_to_search3 = ("TITLE")
                string_to_search4 = ("CHECKED BY")
                string_to_search5 = ("APPROVED BY")
                string_to_search6 = ("STATUS")
                string_to_search7 = ("COMPANY NAME")
                string_to_search8 = ("PAGE")
                string_to_search9 = ("DRAWING NO")
                string_to_search10 = ("CONTRACTOR")
                string_to_search11 = ("LANG")
                string_to_search12 = ("CAD NO")

                newsheet['A1'] = (string_to_search1)    
                newsheet['A2'] = (string_to_search2)
                newsheet['A3'] = (string_to_search3)
                newsheet['A4'] = (string_to_search4)
                newsheet['A5'] = (string_to_search5)
                newsheet['A6'] = (string_to_search6)
                newsheet['A7'] = (string_to_search7)
                newsheet['A8'] = (string_to_search8)
                newsheet['A9'] = (string_to_search9)
                newsheet['A10'] = (string_to_search10)
                newsheet['A11'] = (string_to_search11)
                newsheet['A12'] = (string_to_search12)

                if string_to_search1 in line:
                    newsheet['B1'] = next(read_obj)

                if string_to_search2 in line:
                    newsheet['B2'] = next(read_obj)

                if string_to_search3 in line:
                    newsheet['B3'] = next(read_obj)      

                if string_to_search4 in line:
                    newsheet['B4'] = next(read_obj)

                if string_to_search5 in line:
                    newsheet['B5'] = next(read_obj) 

                if string_to_search6 in line:
                    newsheet['B6'] = next(read_obj)

                if string_to_search7 in line:
                    newsheet['B7'] = next(read_obj)

                if string_to_search8 in line:
                    newsheet['B8'] = next(read_obj)

                if string_to_search9 in line:
                    newsheet['B9'] = next(read_obj)
                
                if string_to_search10 in line:
                    newsheet['B10'] = next(read_obj)
                    
                if string_to_search11 in line:
                    newsheet['B11'] = next(read_obj)
                
                if string_to_search12 in line:
                    newsheet['B12'] = next(read_obj)


                wb.save("DIPassignment.xlsx")  
  
    read_obj.close()
    os.remove("recognized.txt")
    os.remove("DIPimagedetailsWITHOUTspace.txt")
 
def extractTabularData5(file):
    wb = openpyxl.Workbook() 
    sheet = wb.active
    sheet.title = "AMMENDMENTS"
 

    for x in range(1,10):
        for y in range(1,10):
            sheet.cell(row=x, column=y)
    
    im1 = cv2.imread(file, 0)
    im = cv2.imread(file)

 

    blur = cv2.blur(im1,(1,1))

 


    ret,thresh_value = cv2.threshold(blur,100,255,cv2.THRESH_BINARY_INV)

 

    kernel1 = np.ones((4,4), np.uint8)
    img_erosion = cv2.erode(thresh_value, kernel1, iterations=1)

 


    kernel2 = np.ones((4,4),np.uint8)
    dilated_value = cv2.dilate(img_erosion,kernel2,iterations = 1)

 

    contours, hierarchy = cv2.findContours(dilated_value,cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
    cordinates = []
    
            

    for cnt in contours:
        x,y,w,h = cv2.boundingRect(cnt)
        cordinates.append((x,y,w,h))
        cropped = im1[y:y + h, x:x + w] 
        file = open("recognized.txt", "a")
        text = pytesseract.image_to_string(cropped)
        file.write(str(text)) 
        file.write("\n")
        file.close() 
        #bounding the images
        if y > 50:
            cv2.rectangle(im,(x,y),(x+w,y+h),(0,0,255),1)
    
      #Showing and Saving the Detected image
    plt.imshow(im)
    cv2.namedWindow('detecttable', cv2.WINDOW_NORMAL)
    cv2.imwrite('detecttable.jpg',im)
    f = open("recognized.txt", "r")
    content_list = f.readlines()
    for row in range(len(content_list)):
        if(content_list[row] != '\x0c\n'):
         sheet.append([content_list[row]])
    f.close()
    
    
    f1 = open("recognized.txt", "r")
    for line in f1:
        if not line.isspace():
            file2 = open("DIPimagedetailsWITHOUTspace.txt", "a")
            file2.write(line)
            file2.close()
    f1.close()        
            
            
    with open("DIPimagedetailsWITHOUTspace.txt", 'r') as read_obj:
            newsheet =  wb.create_sheet("Extract", 0)
            

            for line in read_obj:    

                string_to_search1 = ("TITLE")
                string_to_search2 = ("CONTRACTOR")
                string_to_search3 = ("DRAWING NUMBER")
                string_to_search4 = ("CHECKED BY")
                string_to_search5 = ("APPROVED BY")
                string_to_search6 = ("STATUS")
                string_to_search7 = ("PROJECT NO")
                string_to_search8 = ("PAGE")
                string_to_search9 = ("UNIT:")
                string_to_search10 = ("CONTRACTOR")
                string_to_search11 = ("LANG")
                string_to_search12 = ("CAD NO")

                newsheet['A1'] = (string_to_search1)    
                newsheet['A2'] = (string_to_search2)
                newsheet['A3'] = (string_to_search3)
                newsheet['A4'] = (string_to_search4)
                newsheet['A5'] = (string_to_search5)
                newsheet['A6'] = (string_to_search6)
                newsheet['A7'] = (string_to_search7)
                newsheet['A8'] = (string_to_search8)
                newsheet['A9'] = (string_to_search9)
                newsheet['A10'] = (string_to_search10)
                newsheet['A11'] = (string_to_search11)
                newsheet['A12'] = (string_to_search12)

                if string_to_search1 in line:
                    newsheet['B1'] = next(read_obj)

                if string_to_search2 in line:
                    newsheet['B2'] = next(read_obj)

                if string_to_search3 in line:
                    newsheet['B3'] = next(read_obj)      

                if string_to_search4 in line:
                    newsheet['B4'] = next(read_obj)

                if string_to_search5 in line:
                    newsheet['B5'] = next(read_obj) 

                if string_to_search6 in line:
                    newsheet['B6'] = next(read_obj)

                if string_to_search7 in line:
                    newsheet['B7'] = next(read_obj)

                if string_to_search8 in line:
                    newsheet['B8'] = next(read_obj)

                if string_to_search9 in line:
                    newsheet['B9'] = next(read_obj)
                
                if string_to_search10 in line:
                    newsheet['B10'] = next(read_obj)
                    
                if string_to_search11 in line:
                    newsheet['B11'] = next(read_obj)
                
                if string_to_search12 in line:
                    newsheet['B12'] = next(read_obj)



    wb.save("DIPassignment.xlsx")  
    os.remove("recognized.txt")
    os.remove("DIPimagedetailsWITHOUTspace.txt")     
    read_obj.close()

def extractTabularData6(file):
    wb = openpyxl.Workbook() 
    sheet = wb.active
    sheet.title = "AMMENDMENTS"
 

    for x in range(1,10):
        for y in range(1,10):
            sheet.cell(row=x, column=y)
    
    im1 = cv2.imread(file, 0)
    im = cv2.imread(file)

 

    blur = cv2.blur(im1,(1,1))

 


    ret,thresh_value = cv2.threshold(blur,100,255,cv2.THRESH_BINARY_INV)

 

    kernel1 = np.ones((4,4), np.uint8)
    img_erosion = cv2.erode(thresh_value, kernel1, iterations=1)

 


    kernel2 = np.ones((4,4),np.uint8)
    dilated_value = cv2.dilate(img_erosion,kernel2,iterations = 1)

 

    contours, hierarchy = cv2.findContours(dilated_value,cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
    cordinates = []
    
            

    for cnt in contours:
        x,y,w,h = cv2.boundingRect(cnt)
        cordinates.append((x,y,w,h))
        cropped = im1[y:y + h, x:x + w] 
        file = open("recognized.txt", "a")
        text = pytesseract.image_to_string(cropped)
        file.write(str(text)) 
        file.write("\n")
        file.close() 
        #bounding the images
        if y > 50:
            cv2.rectangle(im,(x,y),(x+w,y+h),(0,0,255),1)
     
    plt.imshow(im)
    cv2.namedWindow('detecttable', cv2.WINDOW_NORMAL)
    cv2.imwrite('detecttable.jpg',im)
    f = open("recognized.txt", "r")
    content_list = f.readlines()
    for row in range(len(content_list)):
        if(content_list[row] != '\x0c\n'):
         sheet.append([content_list[row]])
    f.close()
    
    
    f1 = open("recognized.txt", "r")
    for line in f1:
        if not line.isspace():
            file2 = open("DIPimagedetailsWITHOUTspace.txt", "a")
            file2.write(line)
            file2.close()
    f1.close()        
            
            
    with open("DIPimagedetailsWITHOUTspace.txt", 'r') as read_obj:
            newsheet =  wb.create_sheet("Extract", 0)
            

            for line in read_obj:    

                string_to_search1 = ("UNIT")
                string_to_search2 = ("DRAWN BY")
                string_to_search3 = ("TITLE")
                string_to_search4 = ("CHECKED BY")
                string_to_search5 = ("APPROVED BY")
                string_to_search6 = ("STATUS")
                string_to_search7 = ("PROJECT NO")
                string_to_search8 = ("PAGE")
                string_to_search9 = ("DRAWING NUMBER")
                string_to_search10 = ("CONTRACTOR")
                string_to_search11 = ("LANG")
                string_to_search12 = ("FONT")

                newsheet['A1'] = (string_to_search1)    
                newsheet['A2'] = (string_to_search2)
                newsheet['A3'] = (string_to_search3)
                newsheet['A4'] = (string_to_search4)
                newsheet['A5'] = (string_to_search5)
                newsheet['A6'] = (string_to_search6)
                newsheet['A7'] = (string_to_search7)
                newsheet['A8'] = (string_to_search8)
                newsheet['A9'] = (string_to_search9)
                newsheet['A10'] = (string_to_search10)
                newsheet['A11'] = (string_to_search11)
                newsheet['A12'] = (string_to_search12)

                if string_to_search1 in line:
                    newsheet['B1'] = next(read_obj)

                if string_to_search2 in line:
                    newsheet['B2'] = next(read_obj)

                if string_to_search3 in line:
                    newsheet['B3'] = next(read_obj)      

                if string_to_search4 in line:
                    newsheet['B4'] = next(read_obj)

                if string_to_search5 in line:
                    newsheet['B5'] = next(read_obj) 

                if string_to_search6 in line:
                    newsheet['B6'] = next(read_obj)

                if string_to_search7 in line:
                    newsheet['B7'] = next(read_obj)

                if string_to_search8 in line:
                    newsheet['B8'] = next(read_obj)

                if string_to_search9 in line:
                    newsheet['B9'] = next(read_obj)
                
                if string_to_search10 in line:
                    newsheet['B10'] = next(read_obj)
                    
                if string_to_search11 in line:
                    newsheet['B11'] = next(read_obj)
                
                if string_to_search12 in line:
                    newsheet['B12'] = next(read_obj)



    wb.save("DIPassignment.xlsx")  
    os.remove("recognized.txt")
    os.remove("DIPimagedetailsWITHOUTspace.txt")    
    read_obj.close() 



    
def assignment1(img_path):
    diagram_segmentation(img_path)
    extractTabularData1(file)
    print("DONE")
def assignment2(img_path):
    diagram_segmentation(img_path)
    extractTabularData2(file)
    print("DONE")
def assignment3(img_path):
    diagram_segmentation(img_path)
    extractTabularData3(file)
    print("DONE")
def assignment4(img_path):
    diagram_segmentation(img_path)
    extractTabularData4(file)
    print("DONE")
def assignment5(img_path):
    diagram_segmentation(img_path)
    extractTabularData5(file)
    print("DONE")
def assignment6(img_path):
    diagram_segmentation(img_path)
    extractTabularData6(file)
    print("DONE")


input = int(input("Enter an image from 1 to 20: "))

while input !=0: 
    if input == 1:    
        assignment1('01.png')
        break
    if input == 2:    
        assignment2('02.png')
        break
    if input == 3:    
        assignment3('03.png')
        break
    if input == 4:    
        assignment4('04.png')
        break
    if input == 5:   
        assignment5('05.png')
        break
    if input == 6:   
        assignment5('06.png')
        break
    if input == 7:     
        assignment5('07.png')
        break
        
    if input == 8:    
        assignment5('08.png')
        break
        
    if input == 9:    
        assignment5('09.png')
        break
        
    if input == 10:    
        assignment5('10.png')
        break
     
    if input == 11:    
        assignment5('11.png')
        break
        
    if input == 12:    
        assignment5('12.png')
        break
        
    if input == 13:    
        assignment5('13.png')
        break
        
    if input == 14:    
        assignment5('14.png')
        break
    
    if input == 15:    
        assignment6('15.png')
        break
        
    if input == 16:    
        assignment6('16.png')
        break
    
    if input == 17:    
        assignment6('17.png')
        break
    
    if input == 18:    
        assignment6('18.png')
        break
        
    if input == 19:    
        assignment6('19.png') 
        break
    
    if input == 20:    
        assignment6('20.png')
        break
        
    else:
        print("Invalid input")
        break
    
    



 